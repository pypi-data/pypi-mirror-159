import sys
import yamlarg
import os
import shutil
import csv
import json
import keyring
from joblib import Parallel, delayed
from datetime import datetime
import textwrap
import yaml
import paramiko
import socket
import asyncio
import keyring.backend

# Set rich to be the default method for printing tracebacks.
from rich.traceback import install

install(show_locals=True)
from rich.console import Console

console = Console()


class LocalKeyring(keyring.backend.KeyringBackend):
    priority = 1

    def set_password(self, servicename, username, password):
        try:
            with open("/keyring.json", "r") as f:
                keyring = json.loads(f.read())
        except:
            keyring = dict()
        if servicename not in keyring:
            keyring[servicename] = dict()
        keyring[servicename][username] = password
        with open("/keyring.json", "w") as f:
            f.write(json.dumps(keyring))

    def get_password(self, servicename, username):
        with open("/keyring.json", "r") as f:
            keyring = json.loads(f.read())
        return keyring[servicename][username]

    def delete_password(self, servicename, username):
        with open("/keyring.json", "r") as f:
            keyring = json.loads(f.read())
        if servicename in keyring:
            keyring[servicename].pop(username, None)
        with open("/keyring.json", "w") as f:
            f.write(json.dumps(keyring))


# Set the keyring backend to a local file if a credential store is not installed.
try:
    keyring.set_password("test", "test", "test")
    keyring.delete_password("test", "test")
except:
    # set the keyring for keyring lib
    keyring.set_keyring(LocalKeyring())


def is_open(ip, port):
    import socket

    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        s.connect((ip, int(port)))
        s.shutdown(2)
        return True
    except:
        return False


def setpass(service, username):
    import keyring
    import getpass

    keyring.set_password(
        service,
        username,
        getpass.getpass("Enter the " + username + " for " + service + ": "),
    )


def get_or_set_password(service, username):
    import keyring

    creds = keyring.get_password(service, username)
    if creds is None:
        setpass(service, username)
        creds = keyring.get_password(service, username)
    return creds


def split_interface(interface):
    try:
        num_index = interface.index(next(x for x in interface if x.isdigit()))
        str_part = interface[:num_index]
        num_part = interface[num_index:]
    except StopIteration:
        return ["", ""]
    return [str_part, num_part]


def get_oui_dict(pkgdir):
    import os

    f_in = open(os.path.join(pkgdir, "templates/wireshark_oui.txt"), "r")
    oui = filter(None, (line.partition("#")[0].rstrip() for line in f_in))
    oui_dict = dict()
    for line in oui:
        part = line.partition("\t")
        if "IEEE Registration Authority" not in part[2]:
            mac_prefix = part[0].replace(":", "").replace(".", "")
            if len(mac_prefix) == 6:
                oui_dict[mac_prefix] = part[2].replace("\t", ", ")
            else:
                if mac_prefix[0:6] not in oui_dict.keys():
                    oui_dict[mac_prefix[0:6]] = dict()
                oui_dict[mac_prefix[0:6]][part[0]] = part[2].replace("\t", ", ")
    return oui_dict


def mac_to_bits(mac_address):
    return int(mac_address.replace(":", "").replace(".", ""), 16)


def mac_subnet(mac_address, subnet):
    mac = mac_to_bits(mac_address)
    low = mac_to_bits(subnet.partition("/")[0])
    high = mac_to_bits(subnet.partition("/")[0]) + int(
        "1" * (48 - int(subnet.partition("/")[2])), 2
    )
    if mac >= low and mac <= high:
        return True
    else:
        return False


def oui_lookup(mac_address, oui_dict):
    mac_address = mac_address.replace(":", "").replace(".", "")
    if mac_address[0:6] in oui_dict.keys():
        if type(oui_dict[mac_address[0:6]]) == str:
            return oui_dict[mac_address[0:6]]
        else:
            for subnet, company in oui_dict[mac_address[0:6]].items():
                if mac_subnet(mac_address, subnet):
                    return company
    else:
        return "(Unknown)"


def test_creds(hostname, username, password):
    import paramiko
    import socket
    import time

    # initialize SSH client
    client = paramiko.SSHClient()
    # add to know hosts
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    try:
        client.connect(
            hostname=hostname, username=username, password=password, timeout=3
        )
        return username, password
    except socket.timeout:
        console.print(
            "Socket timeout connecting to port 22 on " + hostname + ".", style="magenta"
        )
        return None
    except paramiko.AuthenticationException:
        return None
    except paramiko.SSHException:
        time.sleep(60)
        test_creds(hostname, username, password)


def get_device(switch):
    import netmiko.ssh_exception
    from napalm import get_network_driver

    driver = get_network_driver(switch["driver"])
    test_default = True
    while True:
        try:
            if switch["default_creds"] is not None and test_default:
                un = switch["default_creds"]["user"]
                pw = switch["default_creds"]["pass"]
                test_default = False
            else:
                un = get_or_set_password(switch["switch"], "username")
                pw = get_or_set_password(switch["switch"], "password")
            optional_args = {"global_delay_factor": 2, "transport": switch["transport"]}
            if switch["ssh_config"] != "":
                optional_args["ssh_config_file"] = switch["ssh_config"]
            device = driver(switch["switch"], un, pw, optional_args=optional_args)
            device.open()
            break
        except netmiko.ssh_exception.AuthenticationException:
            console.print("Authentication Failed.", style="magenta")
            setpass(switch["switch"], "username")
            setpass(switch["switch"], "password")
    return device


def collect_sw_info(switch):
    from ntc_templates import parse

    device_info = dict()
    # try:
    device = get_device(switch)
    device_info["facts"] = device.get_facts()
    device_info["full-config"] = device.get_config(full=True)
    device_info["config"] = device.get_config()
    device_info["vlans"] = device.get_vlans()
    device_info["arp"] = device.get_arp_table()
    device_info["mac"] = device.get_mac_address_table()
    device_info["users"] = device.get_users()
    device_info["interfaces"] = device.get_interfaces()
    device_info["lldp"] = device.get_lldp_neighbors_detail()
    if switch["driver"] == "ios":
        device_info["cdp"] = device.cli(["show cdp neighbors detail"])[
            "show cdp neighbors detail"
        ]
        device_info["cdp-parsed"] = parse.parse_output(
            "cisco_ios", "show cdp neighbors detail", device_info["cdp"]
        )
        device_info["trans"] = device.cli(["show int trans"])["show int trans"]
        device_info["int"] = device.cli(["show interfaces"])["show interfaces"]
        device_info["int-parsed"] = parse.parse_output(
            "cisco_ios", "show interfaces", device_info["int"]
        )
        device_info["spanning-tree"] = device.cli(["show spanning-tree"])
    if switch["additional_commands"] != "":
        device_info["additional_commands"] = device.cli(
            switch["additional_commands"].split(",")
        )
    print("Information collected from " + switch["switch"])
    return [switch["switch"], device_info]
    # except:
    #     console.print('Data Collection for ' + switch['switch'] + ' failed.', style='red')
    #     return False


def get_switches(filename):
    import csv

    with open(filename, "r") as csvfile:
        switches = csv.DictReader(
            csvfile, fieldnames=["switch", "driver", "transport"], delimiter=","
        )
        next(switches)
        return [switch for switch in switches]


def main():
    cisco_documentation_description = textwrap.dedent(
        """
    Steps to use:
    - Create your switch-list.txt.
    switch-list.txt:
    "
    switch,driver,transport
    192.168.1.1,ios,ssh
    "
    - If you have a large number of switches, create a file with credentials.
    credentials:
    "
    ip,username,password
    192.168.1.1,admin,password
    "
    - To connect with an ssh jumphost, specify --ssh-config = ./ssh_config
    "
    host jumphost
      IdentitiesOnly yes
      IdentityFile ~/.ssh/id_rsa
      User jumphostuser
      HostName ssh.jump.host.domain.tld
      StrictHostKeyChecking no
    
    host * !jumphost
      User pyclass
      # Force usage of this SSH config file
      ProxyCommand ssh -F ./ssh_config -W %h:%p jumphost
      # Alternate solution using netcat
      #ProxyCommand ssh -F ./ssh_config jumphost nc %h %p
      StrictHostKeyChecking no
    
    - Run the cisco-documentation command.
    
    Example Usage:
    cisco-documentation --switch-list switch-list.txt --fetch-info --parse-info --update-excel excel-filename.xlsx
    """
    )
    try:
        pkgdir = sys.modules["cisco_documentation"].__path__[0]
    except KeyError:
        import pathlib

        pkgdir = pathlib.Path(__file__).parent.absolute()

    oui_dict = get_oui_dict(pkgdir)

    args = yamlarg.parse(
        os.path.join(pkgdir, "cisco-documentation.yaml"),
        description=cisco_documentation_description,
    )

    if args["update_wireshark_oui"]:
        import requests

        url = "https://gitlab.com/wireshark/wireshark/raw/master/manuf"
        myfile = requests.get(url)
        open(os.path.join(pkgdir, "templates/wireshark_oui.txt"), "wb").write(
            myfile.content
        )

    if args["load_creds"] != "":
        with open(args["load_creds"], "r") as csvfile:
            creds = csv.DictReader(
                csvfile, fieldnames=["ip", "username", "password"], delimiter=","
            )
            for cred in creds:
                keyring.set_password(cred["ip"], "username", cred["username"])
                keyring.set_password(cred["ip"], "password", cred["password"])
        os.remove(args["load_creds"])

    if args["excel_template"]:
        dstfile = "./Customer City, ST - IP Address Listing.xlsx"
        pkgfile = "templates/Customer City, ST - IP Address Listing.xlsx"
        if not os.path.isfile(dstfile):
            shutil.copy(os.path.join(pkgdir, pkgfile), dstfile)

    if args["fetch_info"]:
        info = dict()
        if args["switch"] != "":
            switch_list = [args["switch"]]
        else:
            switch_list = get_switches(args["switch_list"])
        if args["default_user"] != "" and args["default_pass"] != "":
            default_creds = {"user": args["default_user"], "pass": args["default_pass"]}
        else:
            default_creds = None
        for i in range(len(switch_list)):
            switch_list[i]["default_creds"] = default_creds
            switch_list[i]["ssh_config"] = args["ssh_config"]
            switch_list[i]["additional_commands"] = args["additional_commands"]
        if args["parallel"]:
            results = Parallel(n_jobs=len(switch_list), verbose=0, backend="threading")(
                map(delayed(collect_sw_info), switch_list)
            )
            for result in results:
                try:
                    ip = result[0]
                    device_info = result[1]
                    info[ip] = device_info
                except:
                    pass
        else:
            for switch in switch_list:
                try:
                    ip, device_info = collect_sw_info(switch)
                    info[ip] = device_info
                except:
                    console.print(
                        "Failed to collect information from " + switch["switch"],
                        style="magenta",
                    )
        with open(os.path.join(args["output_dir"], "output.json"), "w") as f:
            f.write(json.dumps(info))

    if args["parse_info"]:
        from napalm.base.helpers import canonical_interface_name

        with open(os.path.join(args["output_dir"], "output.json"), "r") as f:
            info = json.loads(f.read())
        output_dict = dict()
        output_arp = list()
        for sw_ip, device_info in info.items():
            output_dict[sw_ip] = dict()
            # Save config files.
            sw_hostname = device_info["facts"]["hostname"]
            filename = os.path.join(args["output_dir"], sw_ip + "-" + sw_hostname)
            with open(filename + "-running.txt", "w") as f:
                f.write(device_info["config"]["running"])
            with open(filename + "-startup.txt", "w") as f:
                f.write(device_info["config"]["startup"])
            output_dict[sw_ip]["hostname"] = sw_hostname
            output_dict[sw_ip]["interfaces"] = dict()
            for interface, interface_values in device_info["interfaces"].items():
                output_dict[sw_ip]["interfaces"][interface] = dict()
                if interface in device_info["lldp"]:
                    output_dict[sw_ip]["interfaces"][interface]["devices"] = ""
                    output_dict[sw_ip]["interfaces"][interface]["mac"] = [""]
                    output_dict[sw_ip]["interfaces"][interface]["neighbor"] = ""
                    for neighbor in device_info["lldp"][interface]:
                        output_dict[sw_ip]["interfaces"][interface]["neighbor"] += (
                            neighbor["remote_system_name"]
                            + " - "
                            + neighbor["remote_port"]
                        )
                elif 'cdp-parsed' in device_info and interface in [a['local_port'] for a in device_info["cdp-parsed"]]:
                    output_dict[sw_ip]["interfaces"][interface]["devices"] = ""
                    output_dict[sw_ip]["interfaces"][interface]["mac"] = [""]
                    output_dict[sw_ip]["interfaces"][interface]["neighbor"] = ""
                    for neighbor in device_info["cdp-parsed"]:
                        if neighbor['local_port'] == interface:
                            output_dict[sw_ip]["interfaces"][interface]["neighbor"] += (
                                neighbor["destination_host"]
                                + " - "
                                + neighbor["remote_port"]
                            )
                else:
                    output_dict[sw_ip]["interfaces"][interface]["devices"] = len(
                        [
                            True
                            for i in device_info["mac"]
                            if canonical_interface_name(i["interface"]) == interface
                        ]
                    )
                    output_dict[sw_ip]["interfaces"][interface]["mac"] = [
                        i["mac"]
                        for i in device_info["mac"]
                        if canonical_interface_name(i["interface"]) == interface
                    ]
                    if output_dict[sw_ip]["interfaces"][interface]["mac"] == []:
                        output_dict[sw_ip]["interfaces"][interface]["mac"] = [""]
                    output_dict[sw_ip]["interfaces"][interface]["neighbor"] = ""
                output_dict[sw_ip]["interfaces"][interface][
                    "description"
                ] = device_info["interfaces"][interface]["description"]
                output_dict[sw_ip]["interfaces"][interface]["enabled/up"] = (
                    str(device_info["interfaces"][interface]["is_enabled"])
                    + "/"
                    + str(device_info["interfaces"][interface]["is_up"])
                )
                output_dict[sw_ip]["interfaces"][interface]["speed"] = device_info[
                    "interfaces"
                ][interface]["speed"]
                output_dict[sw_ip]["interfaces"][interface]["duplex"] = [
                    i["duplex"]
                    for i in device_info["int-parsed"]
                    if i["interface"] == interface
                ]
                output_dict[sw_ip]["interfaces"][interface]["vlans"] = list()
                for vlan in device_info["vlans"]:
                    if interface in device_info["vlans"][vlan]["interfaces"]:
                        output_dict[sw_ip]["interfaces"][interface]["vlans"].append(
                            vlan
                        )
            for entry in device_info["arp"]:
                if "mac" in entry.keys() and "ip" in entry.keys():
                    output_arp.append(
                        [
                            entry["ip"],
                            entry["mac"].upper().replace(":", "").replace(".", ""),
                            oui_lookup(entry["mac"], oui_dict),
                        ]
                    )
        with open(os.path.join(args["output_dir"], "output.csv"), "w") as f:
            f.write(
                "name\tip\tint\tdevices\tdescription\tenabled/up\tneighbor\tspeed\tduplex\tmac\tvlans\n"
            )
            for sw_ip, device_info in output_dict.items():
                for interface, int_info in device_info["interfaces"].items():
                    device = 0
                    for mac in int_info["mac"]:
                        # name,ip,int,devices,description,enabled/up,neighbor,speed,duplex,mac,vlans
                        output = [
                            device_info["hostname"],
                            sw_ip,
                            interface,
                            str(device),
                            int_info["description"],
                            int_info["enabled/up"],
                            int_info["neighbor"],
                            str(int_info["speed"]),
                            "" if int_info["duplex"] == [] else int_info["duplex"][0],
                            mac.replace(":", ""),
                            ",".join(int_info["vlans"]),
                        ]
                        f.write("\t".join(output) + "\n")
                        device += 1
        with open(os.path.join(args["output_dir"], "arp_output.csv"), "w") as f:
            for entry in output_arp:
                f.write("\t".join(entry) + "\n")

        if args["update_excel"] != "":
            from openpyxl import load_workbook
            from openpyxl.worksheet.table import Table, TableStyleInfo

            wb = load_workbook(args["update_excel"])
            del wb["SWITCHES"]
            ws = wb.create_sheet("SWITCHES")
            ws.append(
                [
                    "",
                    "SWITCH",
                    "SW IP",
                    "INT",
                    "DEVICE",
                    "DESCRIPTION",
                    "LINE PROTO",
                    "NEIGHBOR & PORT",
                    "SPEED",
                    "DUPLEX",
                    "MAC",
                    "VLAN",
                    "IP LOOKUP",
                    "NETWORK",
                    "INTEGRATOR",
                    "DEVICE / APPLICATION",
                    "DEVICE DESCRIPTION",
                    "DEVICE NAME",
                ]
            )

            for sw_ip, device_info in output_dict.items():
                for interface, int_info in device_info["interfaces"].items():
                    device = 0
                    for mac in int_info["mac"]:
                        # name,ip,int,devices,description,enabled/up,neighbor,speed,duplex,mac,vlans
                        output = [
                            "",  # first column is left for navigation links.
                            device_info["hostname"],
                            sw_ip,
                            interface,
                            str(device),
                            int_info["description"],
                            int_info["enabled/up"],
                            int_info["neighbor"],
                            str(int_info["speed"]),
                            "" if int_info["duplex"] == [] else int_info["duplex"][0],
                            mac.replace(":", ""),
                            ",".join(int_info["vlans"]),
                        ]
                        ws.append(output)
                        device += 1
            for row in range(2, ws.max_row):
                ws.cell(
                    row=row,
                    column=13,
                    value="=IFERROR(INDEX(ARP!A:A,MATCH(SWITCHES!K"
                    + str(row)
                    + ',ARP!B:B,0)),"")',
                )
                ws.cell(
                    row=row,
                    column=13,
                    value="=IFERROR(INDEX(OVERVIEW!B:B,MATCH(L"
                    + str(row)
                    + ',OVERVIEW!D:D,0)),"")',
                )
                ws.cell(
                    row=row,
                    column=13,
                    value="=IF(M"
                    + str(row)
                    + '<>"",INDEX(INDIRECT(N'
                    + str(row)
                    + '&"!B1:B99999"),MATCH(M'
                    + str(row)
                    + ",INDIRECT(N"
                    + str(row)
                    + '&"!F1:F99999"),0)),"")',
                )
                ws.cell(
                    row=row,
                    column=13,
                    value="=IF(M"
                    + str(row)
                    + '<>"",INDEX(INDIRECT(N'
                    + str(row)
                    + '&"!C1:C99999"),MATCH(M'
                    + str(row)
                    + ",INDIRECT(N"
                    + str(row)
                    + '&"!F1:F99999"),0)),"")',
                )
                ws.cell(
                    row=row,
                    column=13,
                    value="=IF(M"
                    + str(row)
                    + '<>"",INDEX(INDIRECT(N'
                    + str(row)
                    + '&"!D1:D99999"),MATCH(M'
                    + str(row)
                    + ",INDIRECT(N"
                    + str(row)
                    + '&"!F1:F99999"),0)),"")',
                )
                ws.cell(
                    row=row,
                    column=13,
                    value="=IF(M"
                    + str(row)
                    + '<>"",INDEX(INDIRECT(N'
                    + str(row)
                    + '&"!E1:E99999"),MATCH(M'
                    + str(row)
                    + ",INDIRECT(N"
                    + str(row)
                    + '&"!F1:F99999"),0)),"")',
                )

            ws["A1"].hyperlink = "OVERVIEW!A1"
            ws["A1"].value = "OVERVIEW"
            ws["A1"].style = "Hyperlink"
            ws["A2"].hyperlink = "'OVERVIEW SWITCHES'!A1"
            ws["A2"].value = "OVERVIEW SWITCHES"
            ws["A2"].style = "Hyperlink"
            tab = Table(displayName="SWITCHES", ref="B1:R" + str(ws.max_row))
            style = TableStyleInfo(
                name="Table Style Light 1",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=True,
            )
            tab.tableStyleInfo = style
            ws.add_table(tab)
            col_widths = {
                "A": 120,
                "B": 160,
                "C": 90,
                "D": 123,
                "E": 54,
                "F": 260,
                "G": 84,
                "H": 360,
                "I": 50,
                "J": 65,
                "K": 85,
                "L": 53,
                "M": 90,
                "N": 90,
                "O": 120,
                "P": 150,
                "Q": 250,
                "R": 120,
            }
            for col, width in col_widths.items():
                ws.column_dimensions[col].width = width / 6

            ws = wb["ARP"]
            ws.append([datetime.now()])
            for entry in output_arp:
                ws.append(entry)
            wb.save(args["update_excel"])


def run_commands_cmd(switch):
    device = get_device(switch)
    return [switch["switch"], device.cli(switch["commands"])]


def run_commands():
    run_commands_description = textwrap.dedent(
        """
    This command is used to run multiple commands on a list of switches.
    Switches should be defined in a comma-delimited file with the format of "switch,driver,transport"
    i.e. "192.168.1.1,ios,ssh"
    If the --parallel flag is set, it will run the set of commands on all switches at once.
    Note: configure commands are not supported. You should use 'config-merge' instead.
    
    Example Usage:
    run-commands --parallel --json --command "sh ver~sh span~sh tech"
    
    """
    )
    try:
        pkgdir = sys.modules["cisco_documentation"].__path__[0]
    except KeyError:
        import pathlib

        pkgdir = pathlib.Path(__file__).parent.absolute()

    args = yamlarg.parse(
        os.path.join(pkgdir, "run-commands.yaml"), description=run_commands_description
    )
    switch_list = get_switches(args["switch_list"])
    if args["default_user"] != "" and args["default_pass"] != "":
        default_creds = {"user": args["default_user"], "pass": args["default_pass"]}
    else:
        default_creds = None
    for i in range(len(switch_list)):
        switch_list[i]["default_creds"] = default_creds
        switch_list[i]["ssh_config"] = args["ssh_config"]
        switch_list[i]["additional_commands"] = args["additional_commands"]
    if args["command"] == "":
        command = console.input("Input command to run:")
    else:
        command = args["command"]
    if "~" in command:
        commands = command.split("~")
    else:
        commands = [command]

    if args["parallel"]:
        for switch in switch_list:
            switch["commands"] = commands
        results = Parallel(n_jobs=len(switch_list), verbose=0, backend="threading")(
            map(delayed(run_commands_cmd), switch_list)
        )
        if args["json"]:
            console.print({result[0]: result[1] for result in results})
        else:
            for result in results:
                for k, v in result[1].items():
                    console.print(result[0])
                    console.print(k)
                    console.print(v)
    else:
        for switch in switch_list:
            device = get_device(switch)
            console.print(switch["switch"])
            for command in commands:
                if args["json"]:
                    console.print({"device": switch, "command": device.cli([command])})
                else:
                    console.print(command)
                    console.print(device.cli([command])[command])


def config_merge_cmd(switch):
    if "file" in switch.keys():
        if os.path.isfile(switch["file"]):
            device = get_device(switch)
            device.load_merge_candidate(filename=switch["file"])
            cmp = device.compare_config()
            device.commit_config()
    else:
        device = get_device(switch)
        device.load_merge_candidate(config=switch["text"])
        cmp = device.compare_config()
        device.commit_config()
    return [switch["switch"], cmp]


def config_merge():
    config_merge_description = textwrap.dedent(
        r"""
    This command is used to merge configurations across a list of switches.
    
    If configuration modifications differ for each switch, you can create a directory of configurations.
    ls /config/dir
    1.2.3.4
    1.2.3.5
    
    Example:
    config-merge --config-dir /config/dir --switch-list ./switch-list.txt
    
    If configuration merges will be the same for all switches, you can specify the change with --text.
    
    Example:
    config-merge --config-text "interface fa1/1\nspanning-tree portfast edge\nspanning-tree bpdug en" --switch-list ./switch-list.txt
    
    If config-dir is specified, the configuration within --text will be ignored.
    """
    )
    try:
        pkgdir = sys.modules["cisco_documentation"].__path__[0]
    except KeyError:
        import pathlib

        pkgdir = pathlib.Path(__file__).parent.absolute()

    args = yamlarg.parse(
        os.path.join(pkgdir, "config-merge.yaml"), description=config_merge_description
    )
    switch_list = get_switches(args["switch_list"])
    if args["config_dir"] != "":
        for switch in switch_list:
            switch["file"] = os.path.join(args["config_dir"], switch["switch"])
    elif args["config_text"] != "":
        for switch in switch_list:
            switch["text"] = args["config_text"].replace(r"\n", "\n")
    if args["parallel"]:
        results = Parallel(n_jobs=len(switch_list), verbose=0, backend="threading")(
            map(delayed(config_merge_cmd), switch_list)
        )
        for result in results:
            console.print(result[0])
            console.print(result[1])
    else:
        for switch in switch_list:
            sw, cmp = config_merge_cmd(switch)
            console.print(sw)
            console.print(cmp)


def jinja_merge():
    from jinja2 import Environment

    jinja_merge_description = textwrap.dedent(
        r"""
    This function can be used to template out configurations using a csv file.
    
    Example template:

    {% for section in [global_config, site_specific, switch_specific] %}
      {% for item in global %}
        {% if type(item) == dict %}
          {% for k, v in item.items() %}
          {{ k }}
            {{ v }}
          {% endfor %}
        {% else %}
          {{ item }}
        {% endif %}
    {% endfor %}
    {% endfor %}
    
    Example jinja yaml configuration:
    
    site_specific: # site specific
      - snmp-server community private ro
      - no snmp-server community public ro
    192.168.1.1: # switch specific
      - interface range Fa1/1-8: "switchport mode access\nswitchport access vlan 1\nspanning-tree portfast\nspanning-tree bpdug en"
      - interface Gi1/1: "switchport mode access\ndescription example\nno switchport trunk native vlan\nno switchport trunk allowed vlan"
      - username admin privilege 15 secret password
      - hostname this-is-a-test
    
    For global configuration that span multiple sites, you can specify additional configuration file with --global-config
    
    """
    )
    try:
        pkgdir = sys.modules["cisco_documentation"].__path__[0]
    except KeyError:
        import pathlib

        pkgdir = pathlib.Path(__file__).parent.absolute()

    args = yamlarg.parse(
        os.path.join(pkgdir, "jinja-merge.yaml"), description=jinja_merge_description
    )
    if args["example"]:
        if not os.path.isfile("jinja-config.yaml"):
            shutil.copy(
                os.path.join(pkgdir, "templates/jinja-config.yaml"), "jinja-config.yaml"
            )
        else:
            console.print(
                "Error: Destination file already exists. The file was not overwritten.",
                style="magenta",
            )
    template_file = os.path.join(pkgdir, "templates/jinja.template")
    config_template = Environment().from_string(open(template_file, "r").read())
    if args["global_config"] != "":
        global_config = yaml.load(
            open(args["global_config"], "r"), Loader=yaml.FullLoader
        )
    else:
        global_config = []
    site_specific = yaml.load(open(args["jinja_config"], "r"), Loader=yaml.FullLoader)
    switch_list = get_switches(args["switch_list"])
    switches_to_process = list()
    for switch in switch_list:
        if switch["switch"] in site_specific.keys():
            switch["text"] = config_template.render(
                global_config=global_config,
                site_specific=site_specific["site_specific"],
                switch_specific=site_specific[switch["switch"]],
            )
            switches_to_process.append(switch)
    if args["parallel"]:
        results = Parallel(
            n_jobs=len(switches_to_process), verbose=0, backend="threading"
        )(map(delayed(config_merge_cmd), switches_to_process))
        for result in results:
            console.print(result[0])
            console.print(result[1])
    else:
        for switch in switches_to_process:
            sw, cmp = config_merge_cmd(switch)
            console.print(sw)
            console.print(cmp)


def port_description_cmd(switch):
    device = get_device(switch)
    device.load_merge_candidate(config=switch["text"])
    cmp = device.compare_config()
    device.commit_config()
    return [switch["switch"], cmp]


def port_descriptions():
    port_descriptions_help = textwrap.dedent(
        """
    port-descriptions.txt
    switch,interface,description
    192.168.1.1,Gi1/1,Put your port description here
    
    switch-list.txt
    192.168.1.1,ios,ssh
    
    Example Usage:
    cisco-port-descriptions --apply --parallel
    """
    )
    try:
        pkgdir = sys.modules["cisco_documentation"].__path__[0]
    except KeyError:
        import pathlib

        pkgdir = pathlib.Path(__file__).parent.absolute()

    args = yamlarg.parse(
        os.path.join(pkgdir, "port-descriptions.yaml"),
        description=port_descriptions_help,
    )

    if args["example"]:
        if not os.path.isfile(args["file"]):
            shutil.copy(
                os.path.join(pkgdir, "templates/port-descriptions.csv"), args["file"]
            )
        else:
            console.print(
                "Error: Destination file already exists. The file was not overwritten.",
                style="magenta",
            )
    elif args["apply"]:
        with open(args["file"]) as csvfile:
            reader = csv.DictReader(
                csvfile,
                fieldnames=["switch", "interface", "description"],
                delimiter=",",
            )
            next(reader)
            interface_list = [interface for interface in reader]
        switch_list = get_switches(args["switch_list"])

        config_merge_dict = dict()
        for interface in interface_list:
            if interface["switch"] not in config_merge_dict.keys():
                config_merge_dict[interface["switch"]] = ""
            if interface["description"] != "":
                config_merge_dict[interface["switch"]] += (
                    "interface "
                    + interface["interface"]
                    + "\n description "
                    + interface["description"]
                    + "\n"
                )
            else:
                config_merge_dict[interface["switch"]] += (
                    "interface " + interface["interface"] + "\n no description \n"
                )
        switches_to_update = list()
        for switch in switch_list:
            if switch["switch"] in config_merge_dict.keys():
                switch["text"] = config_merge_dict[switch["switch"]]
                switches_to_update.append(switch)

        if args["parallel"]:
            results = Parallel(
                n_jobs=len(switches_to_update), verbose=0, backend="threading"
            )(map(delayed(port_description_cmd), switches_to_update))
            for result in results:
                console.print(result[0])
                console.print(result[1])
        else:
            for switch in switches_to_update:
                sw, cmp = config_merge_cmd(switch)
                console.print(sw)
                console.print(cmp)


def test_creds(sw, un, pw, timeout=3):
    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    try:
        client.connect(hostname=sw, username=un, password=pw, timeout=timeout)
    # Pass socket.timeout exception back up a level. No need to retry authentication when the host is down / 22 is unavailable.
    # except socket.timeout:
    # console.print('Socket timeout for host ' + sw + '.')
    # return False
    except paramiko.AuthenticationException:
        return False
    # except paramiko.SSHException:
    #     print(f'{BLUE}[*] Quota exceeded, retrying with delay...{RESET}')
    #     # sleep for a minute
    #     time.sleep(60)
    #     return is_ssh_open(hostname, username, password)
    else:
        return True


def ssh_creds():
    ssh_creds_help = textwrap.dedent(
        """
    switch-list.txt
    192.168.1.1
    192.168.1.2

    Example Usage:
    ssh-creds --test --load
    cisco-port-descriptions --apply --parallel
    """
    )
    try:
        pkgdir = sys.modules["cisco_documentation"].__path__[0]
    except KeyError:
        import pathlib

        pkgdir = pathlib.Path(__file__).parent.absolute()
    args = yamlarg.parse(
        os.path.join(pkgdir, "ssh-creds.yaml"), description=ssh_creds_help
    )
    with open(args["switch_list"], "r") as f:
        switches = [line.strip("\n") for line in f.readlines()]
    if args["default_user"] == "":
        default_user = console.input("Input the default username: ")
    else:
        default_user = args["default_user"]
    if args["default_pass"] == "":
        default_pass = console.input("Input the default password: ", password=True)
    else:
        default_pass = args["default_password"]
    creds = dict()
    for switch in switches:
        try:
            if test_creds(switch, default_user, default_pass):
                creds[switch] = {"user": default_user, "pass": default_pass}
        except paramiko.AuthenticationException:
            while switch not in creds.keys():
                un = console.input("Username: ")
                pw = console.input("Password: ", password=True)
                try:
                    test_creds(switch, un, pw)
                    creds[switch] = {"user": un, "pass": pw}
                except paramiko.AuthenticationException:
                    console.print("Authentication failed.")
                except paramiko.SSHException:
                    console.print("Authentication quota exceeded.")
        except socket.timeout:
            console.print("Socket timeout connecting to " + switch + ". Skipping.")
        except paramiko.ssh_exception.NoValidConnectionsError:
            console.print("Error connecting to " + switch + ". Skipping.")
    if args["save_creds"]:
        with open(args["filename"], "w") as f:
            for sw in creds.keys():
                f.write([sw, creds[sw]["user"], creds[sw]["pass"]].join(",") + "\n")
