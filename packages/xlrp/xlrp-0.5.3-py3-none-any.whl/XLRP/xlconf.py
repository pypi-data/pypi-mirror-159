from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image, PILImage
import win32gui, win32print, win32con


class _XlConf:
    """
    进行保存图片的excel、sheet的基础格式设置和保存plot图像的类
    """
    def __init__(self, filepath, sheetname:str, pic_list=None, pic_width=(500, 500, 500)):
        self.filepath = filepath
        self.pic_list = pic_list
        self.pic_width = pic_width
        self.sheet = sheetname
        self.dpi = win32print.GetDeviceCaps(win32gui.GetDC(0), win32con.LOGPIXELSX)     # 获取当前设备分辨率

    def set_cell_size(self, cells=('A1', 'B1', 'A2'), create_new=False):
        """
        设置单元格格式
        :param cells: 需要进行操作的单元格
        :param create_new: 是否需要创建一个新的sheet
        :return:
        """
        wb = load_workbook(self.filepath)
        if create_new is True:
            ws = wb.create_sheet(self.sheet)
        else:
            ws = wb[self.sheet]
        for index, pic in enumerate(self.pic_list):             # 获取所有的图片
            w, h = PILImage.open(pic).size                      # 获取图片的尺寸
            size_ratio = self.pic_width[index] / w              # 获取需要设置的大小和图片原始尺寸的宽度比例
            cell_width = (self.pic_width[index] / self.dpi * 25.4) / 1.8    # 像素转换英寸，设置为单元格大小
            cell_col = cells[index][0]
            cell_row = cells[index][1]
            ws.row_dimensions[cell_row].height = h * size_ratio         # 设置单元格的大小，按图片的尺寸来进行设置
            ws.column_dimensions[cell_col].width = cell_width
        wb.save(self.filepath)
        wb.close()
        return self

    def save_plot(self, cells=('A1', 'B1', 'A2:B2')):
        """
        保存图片到excel的方法，最后一个占用两个单元格
        :param cells: 需要保存进入的单元格
        :return:
        """
        wb = load_workbook(self.filepath)
        ws = wb[self.sheet]
        for index, pic in enumerate(self.pic_list):
            w, h = PILImage.open(pic).size
            ratio = self.pic_width[index] / w
            img = Image(pic)            # 将图片使用openpyxl的Image读取出来，作为写入的资源
            if index == 2:
                img.width, img.height = w * ratio * 2, h * ratio * 2        # 根据缩放比对图片进行等比的缩放操作
                ws.merge_cells(cells[index])                # 如果是第三个参数的单元格，则进行合并操作
                cell = cells[index].split(':')[0]
            else:
                img.width, img.height = w * ratio, h * ratio
                cell = cells[index]
            align = Alignment(horizontal='center', vertical='center')       # 设置单元格居中对齐
            ws[cell].alignment = align
            ws.add_image(img, cell)             # 将图片放入excel中保存
        wb.save(self.filepath)
        wb.close()




