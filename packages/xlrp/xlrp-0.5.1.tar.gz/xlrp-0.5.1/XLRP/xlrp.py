import time
import traceback
from functools import wraps, reduce
from XLRP.drawplot import _DrawPlot
from pathlib import Path
import os
from XLRP.xlconf import _XlConf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from XLRP.xl_ddt import *
import sys
import threading
import inspect
import glob
import platform
from XLRP.xl_logger import Logger


_test_result = {"sys": '', "model": [], 'step': []}     # 用例收集全局变量
_error = {}             # 收集错误信息变量
_passed_count = 0       # 通过总数
_failed_count = 0       # 失败总数
_error_count = 0        # 错误总数
CACHE_PATH = Path().cwd() / '.xlrp_cache'       # xlrp运行总文件夹
CACHE_PATH.mkdir(exist_ok=True)
_XL_LOCK = threading.RLock()
file_mark = None                # 文件多进程标识
log_file = None                 # 日志文件路径
_old_model = None               # model每次赋值，为当前所属模块名称
_ALREADY_FUNC = []              # 定义一个存放已运行函数的列表，作为run_class中的过滤
_RUNNER_COUNT = 0               # run模块运行次数，主要为了使用run_discover调用run_class后的显示效果
_FIRST_RUN = True               # 是否第一次运行，也是为了显示上述一致的显示效果


def _log_path():
    """
    创建日志文件代码
    :return: 日志文件地址
    """
    path_dir = CACHE_PATH / 'logs'
    path_dir.mkdir(exist_ok=True)
    file = path_dir / f'xlrp{file_mark}.log'
    file.touch(exist_ok=True)
    file_path = os.fspath(file)
    return file_path

def _write_log(file_path, msg):
    """
    写入日志文件代码
    :param file_path: 日志文件地址
    :param msg: 日志文件写入内容
    :return:
    """
    with open(file_path, 'a', encoding='utf8') as f:
        f.writelines(msg)
        f.write('\n')
        f.close()

class SysName:
    """
    获取系统名称的类
    """
    def __init__(self, sys_name):
        self.sys_name = sys_name

    def __enter__(self):
        pass

    def __exit__(self, exc_type, exc_val, exc_tb):
        _test_result["sys"] = self.sys_name             # 退出with块时收集系统名称

    def __call__(self, cls):
        _test_result['sys'] = self.sys_name             # 装饰器运行收集系统名称

        @wraps(cls)
        def decorator(*args, **kwargs):
            try:
                res = cls(*args, **kwargs)
                return res
            except:
                ...
        return decorator

class ModelName:
    """
    获取模块名称类
    """
    def __init__(self, model_name):
        self.model = model_name

    def __enter__(self):
        global _old_model
        _old_model = self.model                     # 进入with块为old_model全局变量赋值，模块名在步骤中添加

    def __exit__(self, exc_type, exc_val, exc_tb):
        ...

    def __call__(self, func):               # 装饰器方法
        @wraps(func)
        def decorator(*args, **kwargs):
            global _old_model
            _old_model = self.model
            # _test_result["model"].append(self.model)
            try:
                res = func(*args, **kwargs)
                return res
            except:
                ...
        return decorator

class StepName:
    """
    获取step步骤(用例)的类
    """
    def __init__(self, step_name):
        self.step = step_name

    def __enter__(self):
        self.start_time = time.time()

    def __exit__(self, exc_type, exc_val, exc_tb):          # with块退出时运行逻辑
        global _failed_count, _passed_count, _error_count, _test_result, _old_model, _FUNC_INFO

        # print(sys._getframe(1).f_lineno, sys._getframe(1).f_code.co_filename, sys._getframe(1).f_code.co_name)
        func_file = sys._getframe(1).f_code.co_filename
        func_line = sys._getframe(1).f_lineno
        func_name = sys._getframe(1).f_code.co_name
        log = f'File: "{func_file}:{func_line}" FuncName: {func_name} Status: '
        if any([exc_type, exc_val, exc_tb]):        # 当报错信息有任意一个不为False时，获取并进行写入
            # error_str = str([''.join(x) for x in traceback.format_exception(exc_type, exc_val, exc_tb)][0])
            error_str = reduce(lambda x, y: x + y, traceback.format_exception(exc_type, exc_val, exc_tb))
            _error[self.step] = error_str
            if exc_type is AssertionError:                  # 如果错误的类型是断言错误，则设置用例运行为False状态
                case_status = False
                # Logger.error(f'{self.step}----Failed')
                Logger(log_file).error(log + f'{self.step}----Failed')
                _failed_count += 1
            else:
                case_status = 'Error'                       # 不为断言错误则视为程序运行出错，给Error状态
                # Logger.error(f'{self.step}----Error')
                Logger(log_file).error(log + f'{self.step}----Error')
                _error_count += 1
            e = traceback.format_exception(exc_type, exc_val, exc_tb)
            _write_log(log_file, e)         # 写入日志文件
            if type(e) is list:
                for msg in e:
                    print(msg, end='')
            else:
                print(e)
            print()
        else:
            _passed_count += 1              # 没有报错信息的时候则视为通过用例
            case_status = True
            error_str = None
            Logger(log_file).success(log + f'{self.step}---Passed')
        stop_time = time.time()
        run_time = round(stop_time - self.start_time, 3)        # 用例运行时间取小数点后3位
        # 添加用例步骤信息和模块名称信息，模块名称为赋值的_old_model
        _test_result["step"].append({self.step: {'run_time': run_time, 'status': case_status, 'msg': error_str}})
        _test_result['model'].append(_old_model)

    def __call__(self, func):
        step = self.step

        func_name = sys._getframe(1).f_code.co_name
        func_file = sys._getframe(1).f_code.co_filename
        func_line = sys._getframe(1).f_lineno
        @wraps(func)
        def decorator(*args, **kwargs):
            global _passed_count, _failed_count, _error_count, _test_result, _old_model
            log = f'File: "{func_file}:{func_line}" FuncName: {func_name} Status: '
            start_time = time.time()
            try:
                res = func(*args, **kwargs)
                if type(res) is bool:           # 当返回的内容是布尔值的时候，则作为用例判断标准
                    case_status = res
                    if case_status is False:
                        log = log + f'{step}---Failed'
                        _failed_count += 1
                        # Logger.error(log)
                        Logger(log_file).success(log)
                        _write_log(log_file, f'The custom assertion failed, function {func.__name__} return {res}')
                        print(f'The custom assertion failed, function {func.__name__} return {res}\n')
                    else:
                        log = log + f'{step}----Passed'
                        Logger(log_file).success(log)

                        _passed_count += 1
                else:
                    case_status = True
                    log = log + f'{step}----Passed'
                    Logger(log_file).success(log)
                    _passed_count += 1
                error_str = None
            except Exception as e:
                case_status = False
                res = None
                # error_str = str([''.join(x) for x in traceback.format_exc()][0])
                error_str = reduce(lambda x, y: x + y, traceback.format_exc())      # 将所有错误信息写入拼接成一个字符串
                exc_type, exc_val, exc_tb = sys.exc_info()          # 再次获取一次错误信息
                if exc_type is AssertionError:          # 判断是否为断言错误，不是断言错误则给定Error
                    log = log + f'{step}----Failed'
                    _failed_count += 1
                else:
                    log = log + f'{step}----Error'
                    _error_count += 1
                Logger(log_file).error(log)
                e = traceback.format_exc()
                _write_log(log_file, e)
                if type(e) is list:
                    for msg in e:
                        print(msg)
                else:
                    print(e)
                print()
            stop_time = time.time()
            run_time = round(stop_time - start_time, 3)         # 用例运行时间给定小数点后三位
            # 添加模块名称和步骤信息
            _test_result['model'].append(_old_model)
            _test_result["step"].append({step: {"run_time": run_time, "status": case_status, 'msg': error_str}})
            return res
        return decorator

class _ProcessingData:
    """
    进行用例数据处理，作为绘图程序数据
    """
    def __init__(self, data_list: dict):
        self.data_list = data_list
        self.model_list = self.data_list['model']
        self.filter_models = []
        [self.filter_models.append(x) for x in self.model_list if x not in self.filter_models]
        self.steps = self.data_list['step']
        self.zips_data = list(zip(self.model_list, self.steps))

    def pie_data(self):
        """
        修整饼图数据
        :return:
        """
        passed = 0
        failed = 0
        error = 0
        for dc in self.steps:
            status = list(dc.values())[0]['status']     # 获取_test_result的step中的状态，根据状态计算用例失败成功错误总数
            if status is True:
                passed += 1
            elif status == 'Error':
                error += 1
            elif status is False:
                failed += 1
            else:
                failed += 1
        return passed, failed, error

    def bar_data(self):
        """
        修整柱状图的测试数据
        :return:
        """
        dc_data = {}
        for model in self.filter_models:
            dc_data[model] = {'通过': 0, '失败': 0, '错误': 0}
            for case in self.zips_data:
                if model == case[0]:
                    status = list(case[1].values())[0]['status']
                    if status is True:
                        dc_data[model]['通过'] += 1
                    elif status is False:
                        dc_data[model]['失败'] += 1
                    elif status == 'Error':
                        dc_data[model]['错误'] += 1
                    else:
                        dc_data[model]['失败'] += 1
        return dc_data

    def plot_data(self):
        """
        修正绘制折线图的测试数据
        :return:
        """
        dc_data = {}
        for model in self.filter_models:
            dc_data[model] = {}
            for case in self.zips_data:
                if model == case[0]:
                    run_time = {list(case[1].keys())[0]: list(case[1].values())[0]["run_time"]}
                    dc_data[model].update(run_time)
        return dc_data

class Runner:
    def __init__(self, show_plot=False, file_mark=''):
        self.show_plot = show_plot              # 布尔值，为True则程序运行立即显示图像
        self.save_path = CACHE_PATH / 'plot'
        self.save_path.mkdir(exist_ok=True)
        self.save_path = os.fspath(self.save_path)
        self.file_mark = file_mark              # 文件标识，用于多进程为文件命名

    def run(self, func, param_iter=None):
        """
        运行单个用例的方法
        :param func: 用例函数或者方法名
        :param param_iter: 参数，使用同parameter，需要列表或者元组形式，会循环传入到函数或者方法中
        :return: self
        """
        global file_mark, log_file
        if self.file_mark:
            file_mark = '_' + self.file_mark
        else:
            file_mark = self.file_mark          # 获取程序file_mark标识
        log_file = _log_path()                  # 生成并获取日志文件
        start_time = time.strftime('%Y/%m/%d-%H:%M:%S')
        start_label = '=' * 30 + start_time + '  开始运行' + '=' * 30
        _write_log(file_path=log_file, msg=start_label)
        print('='*30 + ' 开始测试 ' + '='*30)
        print('-'*20 + f' 开始用例{func.__name__} ' + '-'*20)
        _write_log(log_file, '-'*20 + f' 开始用例{func.__name__} ' + '-'*20)
        # 判断是否有被传入param_iter，如果有传入，则进行循环取值并放入func函数中去运行
        if type(param_iter) not in (list, tuple):
            if param_iter is None:
                try:
                    func()
                except:
                    ...
            else:
                raise ValueError("The parameters must be list, tuple, dict or None")
        else:
            for param in param_iter:
                print(f'\033[1;50;34mparams: {param}\033[0m')
                _write_log(log_file, f'params: {param}')
                try:
                    if type(param) in (list, tuple):
                        func(*param)
                    elif type(param) is dict:
                        func(**param)
                    else:
                        func(param)
                except:
                    ...
        print('-'*20 + f' 结束用例{func.__name__} ' + '-'*20 + '\n')
        print('='*30, f'Passed: {_passed_count} Failed: {_failed_count} Error: {_error_count}', '='*30)
        _write_log(log_file, '-'*20 + f' 结束用例{func.__name__} ' + '-'*20 + '\n')
        _write_log(log_file, '='*30 + f'Passed: {_passed_count} Failed: {_failed_count} Error: {_error_count}' + '='*30)
        end_time = time.strftime('%Y/%m/%d-%H:%M:%S')
        end_label = '='*30 + end_time + '  结束run运行' + '=' * 30
        _write_log(log_file, end_label + '\n')
        return self

    def run_class(self, cls):
        """
        运行class的方法
        :param cls: 需要运行的class实例
        :return:
        """
        global file_mark, log_file, _ALREADY_FUNC, _RUNNER_COUNT, _FIRST_RUN
        _RUNNER_COUNT -= 1                      # 使用此变量，控制结束测试的显示点，为<=0时显示
        if self.file_mark:
            file_mark = '_' + self.file_mark
        else:
            file_mark = self.file_mark              # 判断是否有传入文件标识
        log_file = _log_path()
        # log_conf.deep = 5
        # log_conf.path = log_file
        # log_conf.module = ['success', 'error', 'warning', 'tip']
        # print('user_func', USE_FUNC)
        if _FIRST_RUN is True:
            start_time = time.strftime('%Y/%m/%d-%H:%M:%S')
            start_label = '=' * 30 + start_time + '  开始运行' + '=' * 30
            _write_log(file_path=log_file, msg=start_label)
            print('='*50 + ' 开始测试 ' + '='*50)
        func_names = [x.__name__ for x in FUNC_NAME]                # 获取class中添加了xl_data装饰器的方法
        # 获取cls类中，不存在于func_names的所有用例
        all_func = [x for x in cls.__dir__() if x.startswith('xl') and (x not in func_names)]
        user_func_key = [list(x.keys())[0] for x in USE_FUNC]            # 需要被运行的方法和参数，拿到所有方法名称
        runner_list = list(set(all_func).difference(set(user_func_key)))    # 取出所有cls所有方法和需要运行的方法的不同值
        # runner_list.sort()
        # print(runner_list)
        [USE_FUNC.append({x: ()}) for x in runner_list]         # 将不同值添加进入USE_FUNC中，未被xl_data参数的的默认没有参数
        USE_FUNC.sort(key=lambda x: list(x.keys())[0])          # 对USE_FUNC进行排序
        # 循环所有的方法，传入参数并运行
        for data in USE_FUNC:
            if data in _ALREADY_FUNC:                       # 每次运行将已运行添加进入列表，如果已运行，则直接跳过
                continue                                    # 跳过主要是为了run_discover避免重复数据出现的情况
            func = list(data.keys())[0]
            param = list(data.values())[0]
            print('-'*20 + f' 开始用例{func} ' + '-'*20)
            _write_log(log_file, '-'*20 + f' 开始用例{func} ' + '-'*20)
            try:
                _write_log(log_file, f'params: {param}')
                print(f'\033[1;50;34mparams: {param}\033[0m')
                # func = list(data.keys())[0]
                # param = list(data.values())[0]
                if type(param) in (list, tuple):
                    eval(f'cls.{func}(*{param})')           # 使用eval构建运行
                elif type(param) is dict:
                    eval(f'cls.{func}(**{param})')
                else:
                    eval(f'cls.{func}(*{[param]})')         # 构建成列表形式解包传参
            except ValueError:
                # traceback.print_exc()
                ...
            print('-'*20 + f' 结束用例{func} ' + '-'*20 + '\n')
            _write_log(log_file, '-'*20 + f' 结束用例{func} ' + '-'*20 + '\n')
            _ALREADY_FUNC.append(data)
        if _RUNNER_COUNT <= 0:
            print('=' * 30, f'Passed: {_passed_count} Failed: {_failed_count} Error: {_error_count}', '=' * 30)
            _write_log(log_file, '=' * 30 + f'Passed: {_passed_count} Failed: {_failed_count} Error: {_error_count}' + '=' * 30)
            end_time = time.strftime('%Y/%m/%d-%H:%M:%S')
            end_label = '='*30 + end_time + '  结束class运行' + '=' * 30
            _write_log(log_file, end_label + '\n')
        _FIRST_RUN = False              # 代码运行一次过后，设置为False，声明可能是循环运行，不再打印开始测试提示
        return self

    def run_discover(self, start_dir, sys_name, patten='xl*.py', cls_start_mark='Xl'):
        """
        收集所有文件夹中的用例
        :param start_dir: 用例存放的文件夹
        :param sys_name: 系统名称
        :param patten: 收集何种格式的py文件
        :param cls_start_mark: 标记收集以什么开头的类名
        :return:
        """
        global _test_result, _RUNNER_COUNT
        current_sys = platform.system()             # 获取当前系统，是windows则用\作为分隔符
        split_mark = '/'
        if current_sys == 'Windows':
            split_mark = '\\'
        ab_path = os.path.abspath(start_dir)        # 将用户输入的路径转换为绝对路径
        if ab_path not in sys.path:                 # 如果此路径不在sys.path，python导入路径，则添加进去
            sys.path.insert(0, ab_path)
        case_files = glob.glob(ab_path + split_mark + patten)       # 获取所有patten指定的路径下py文件
        # 取出所有的py文件名称，去除路径和文件扩展名
        module_name = [x.split(split_mark)[-1].split('.')[0] for x in case_files]
        module_name.sort()              # 模块名称进行简单排序操作
        modules = []
        try:
            for name in module_name:                # 循环所有模块名称，进行import导入，获取module名称
                modules.append(__import__(name))
        except ModuleNotFoundError as e:
            traceback.print_exc()
        all_cls = []
        all_funcs = []
        for m in modules:               # 循环所有module，找出class类和function函数
            cls = inspect.getmembers(inspect.getmodule(m), inspect.isclass)
            funcs = inspect.getmembers(inspect.getmodule(m), inspect.isfunction)
            all_cls += cls
            all_funcs += funcs
        # print(all_cls)
        # 过滤类名，将可运行类存入module_class列表中
        module_class = [j for x, j in all_cls if j.__module__.split('.')[-1] in module_name]
        # 在使用xl_ddt会把class变成function，当未收集到类时，收集所有function
        if module_class == []:
            funcs_name = [x[0] for x in all_funcs]
            if 'xl_ddt' in funcs_name:                      # 如果xl_ddt在收集到的function中
                for func in all_funcs:                      # 循环所有的function，使用cls_start_mark进行过滤，取出class
                    if func[0].startswith(cls_start_mark):
                        module_class.append(func[1])
            else:
                raise ModuleNotFoundError('No class were collected, please check cls_start_mark\n'
                                          '未收集到class，请检查cls_start_mark参数')
        if module_class == []:          # 如果上述操作都未能找到，则抛出错误
            raise ModuleNotFoundError('No class were collected, please check cls_start_mark\n'
                                      '未收集到class，请检查cls_start_mark参数')
        module_class = [x for x in module_class if x.__name__.startswith(cls_start_mark)]       # 根据cls_start_mark过滤
        _RUNNER_COUNT = len(module_class)
        for case_class in module_class:                             # 循环所有模块内的class
            # if case_class.__name__.startswith(cls_start_mark):      # 再次进行cls_start_mark进行过滤，满足要求则循环运行
            self.run_class(case_class())
        _test_result['sys'] = sys_name                  # 运行完成重新将sys_name赋值给用例结果的sys
        return self

    def plot_save_excel(self, file_path=None, width_list=(700, 700, 700)):
        """
        将plot图片保存到excel的代码
        :param file_path: excel文件路径
        :param width_list: 图片的宽度(默认第三个数据，会在原基础上X2，改动在xlconf文件中)
        :return: self
        """
        sys_name = _test_result.get('sys')          # 获取系统名称
        prd = _ProcessingData(_test_result)         # 获取各个绘图程序的数据
        pie_data = prd.pie_data()                   # 饼图数据
        bar_data = prd.bar_data()                   # 柱状图数据
        plot_data = prd.plot_data()                 # 折线图数据
        # 进行图像绘制
        dp = _DrawPlot(sys_name=sys_name,file_mark=file_mark, save_path=self.save_path)
        pie_path = dp.draw_pie(通过=pie_data[0], 失败=pie_data[1], 错误=pie_data[2], show_plot=self.show_plot)
        bar_path = dp.draw_bar(**bar_data, show_plot=self.show_plot)
        plot_path = dp.draw_plot(**plot_data, show_plot=self.show_plot)
        pic_list = [pie_path, bar_path, plot_path]
        # 如果传入了excel文件路径，则将图片进行保存
        if file_path is not None:
            # dp.plot_to_excel(file_path=file_path, pic_list=pic_list, width_list=width_list, cells=cells)
            xc = _XlConf(filepath=file_path, sheetname='TestReport', pic_list=pic_list, pic_width=width_list)
            xc.set_cell_size(create_new=True)
            del xc
            xc2 = _XlConf(filepath=file_path, sheetname='TestReport', pic_list=pic_list, pic_width=width_list)
            xc2.save_plot()

    def save_default(self):
        """
        创建excel保存用例数据，并新建sheet存放报告图表
        :return:
        """
        global CACHE_PATH
        sys_name = _test_result.get('sys')                  # 获取系统名称
        excel_name = f'{sys_name}EXCEL报告{file_mark}.xlsx'   # 构建测试报告
        report_path = CACHE_PATH / 'report'                 # 创建报告路径
        report_path.mkdir(exist_ok=True)
        model_list = _test_result['model']              # 获取model列表
        steps = _test_result['step']                    # 获取step用例列表
        # excel样式设置
        wb = openpyxl.Workbook()
        ws = wb.create_sheet('TestCases')
        title_font = Font(size=16, name='微软雅黑', bold=True)
        label_font = Font(size=14, name='微软雅黑')
        content_font = Font(size=11, name='微软雅黑', bold=True)
        false_font = Font(size=11, name='微软雅黑', color='cc3333', bold=True)
        title_fill = PatternFill(fgColor='adc2eb', fill_type='solid')   # 填充单元格样式
        label_fill = PatternFill(fgColor='b3b3b3', fill_type='solid')
        # false_fill = PatternFill(fgColor='cc3333', fill_type='solid')
        cell_align = Alignment(horizontal='center', vertical='center', wrap_text=True)  # 对齐样式，自动换行
        ws.merge_cells('A1:F1')         # 第一行title设置合并
        ws['A1'].value = f'{sys_name}系统运行用例情况表'
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = cell_align
        labels = ['编号', '模块', '用例名称', '运行时间', '运行结果', '异常信息']
        # excel用例存放
        for index, label in enumerate(labels):                  # 设置表格第二行的标签
            label_cell = ws.cell(row=2, column=index + 1)
            label_cell.value = label
            label_cell.font = label_font
            label_cell.fill = label_fill
            label_cell.alignment = cell_align
        for index, case_data in enumerate(zip(model_list, steps)):
            model = case_data[0]
            step = case_data[1]
            case_name = list(step.keys())[0]
            runtime = step[case_name]['run_time']
            status_origin = step[case_name]['status']
            if status_origin is True:
                status = 'PASSED'
            elif status_origin is False:
                status = 'FAILED'
            else:
                status = 'ERROR'
            msg = step[case_name]['msg']
            data = [str(index + 1), model, case_name, runtime, status, msg]
            for col, param in enumerate(data):
                case_cell = ws.cell(row=index + 3, column=col + 1)
                case_cell.value = param
                if status != 'PASSED':
                    case_cell.font = false_font
                else:
                    case_cell.font = content_font
                case_cell.alignment = cell_align
                # if status is not True:
                #     case_cell.fill = false_fill
        sava_path = os.fspath(report_path) + '/' + excel_name
        wb.save(sava_path)
        wb.close()
        del wb
        wb2 = openpyxl.load_workbook(sava_path)
        ws2 = wb2['TestCases']
        try:
            wb2.remove(wb2['Sheet'])            # 移除创建excel会多余的第一个sheet页面
        except:
            ...
        for col in range(1, len(labels) + 1):
            width = 25
            if col == len(labels):
                width = 50
            ws2.column_dimensions[get_column_letter(col)].width = width
        old_model = ''
        # 设置相同模块合并
        for row in range(3, ws2.max_row + 1):
            current_model = ws2.cell(row=row, column=2).value
            if old_model == current_model:
                ws2.merge_cells(f'B{row - 1}:B{row}')
            else:
                old_model = current_model
        wb2.save(sava_path)
        wb2.close()
        del wb2
        self.plot_save_excel(file_path=sava_path)
