import copy
import random
import math
from concurrent.futures.thread import ThreadPoolExecutor

import os
import json
import shutil
import yaml
import pandas as pd
import openpyxl as xl
from openpyxl.styles import Font, Alignment
from pypinyin import pinyin, NORMAL
import xml.etree.ElementTree as ET
import cv2
import math
import numpy as np


IMG_TYPES = ['jpg', 'png', 'JPG', 'PNG']

def create_empty_json_instance(img_file_path: str):
    '''
    :param img_file_path: img路径
    :return: 构建一个空的labelme json instance对象
    '''
    instance = {'version': '1.0',
                'shapes': [],
                'imageData': None,
                'imagePath': img_file_path[img_file_path.rindex(os.sep)+1:]}
    img = cv2.imread(img_file_path)
    instance['imageHeight'], instance['imageWidth'], instance['imageDepth'] = img.shape
    # instance_to_json(instance, img_file_path[:img_file_path.rindex('.')]+'.json')
    return instance

def json_to_instance(json_file_path):
    with open(json_file_path, 'r', encoding='utf-8') as f:
        instance = json.load(f)
    return instance

def instance_to_json(instance, json_file_path):
    with open(json_file_path, 'w', encoding='utf-8') as f:
        content = json.dumps(instance, ensure_ascii=False, indent=2)
        f.write(content)

def yaml_to_instance(yaml_file_path):
    """
    yaml_file_path: yaml文件路径
    return yaml_instance
    """
    with open(yaml_file_path, 'r', encoding='utf-8') as f:
        config = f.read()
    cfg = yaml.load(config, yaml.FullLoader)
    return cfg

def filtrate_file(path):
    list = os.listdir(path)
    for obj in list:
        file_path = os.path.join(path, obj)
        if not os.path.isfile(file_path) or obj[obj.rindex('.') + 1:] not in ['json', 'jpg', 'png']: continue

def word_to_pinyin(word):
    """
    @param word:
    @return:
    """
    # pinyin return [[py1],[py2],...,[pyn]]
    s = ''
    for i in pinyin(word, style=NORMAL):
        s += i[0].strip()
    return s

def read_txt(path):
    with open(path, "r", encoding='utf-8') as f:  # 打开文件
        data = f.readlines()  # 读取文件
    return data

# 写入新的excel   ### 内容， 保存路径, 路径下sheet， 表头， 表头列， 插入位置行数，插入位置列数
def content_to_excel(content, save_path, sheet_name = None, header = False, index=False, row=None, col=None):
    excel_data = pd.DataFrame(content)
    writer = pd.ExcelWriter(save_path)  # 写入Excel文件
    excel_data.to_excel(writer, sheet_name=sheet_name, header=header, index=index, startrow=row, startcol=col)
    writer.save()
    writer.close()

# 追加写入excel（可指定位置）
def write_excel_xlsx_append(file_path, data_name, data, row=0, col=0, sheet_name='sheet1'):
    workbook = xl.load_workbook(file_path)  # 打开工作簿
    sheet = workbook[sheet_name]
    for i in range(0, len(data)):
        for j in range(0, len(data[i])):
            sheet.cell(row=(i + row + 1), column=(j + col + 1), value=data[i][j])  # 追加写入数据，注意是从i+row行，j + col列开始写入
    workbook.save(file_path)  # 保存工作簿
    print("xlsx格式表格【追加】写入{}成功！".format(data_name))

# 创建一个excel，sheet
def create_empty_excel(save_path, sheet):
    wb = xl.Workbook()
    ws = wb.create_sheet(sheet, 0)
    wb.save(save_path)

# 读取excel中的sheet
def read_excel(excel_path, sheet_name):
    book = xl.load_workbook(excel_path)
    sheet = book[sheet_name]
    return sheet

# 美化excel
def beautify_excel(excel_path):
    font = Font(name='宋体', size=11, color='FF000000', bold=True, italic=False)
    align = Alignment(horizontal='center', vertical='center', wrap_text=False)

    book = xl.load_workbook(excel_path)  # 加载excel
    sheet_names = book.sheetnames  # 获取所有的sheet名称
    sheet_names.remove('Sheet')

    for sheet in sheet_names:
        for row in book['{}'.format(sheet)]['A1:A5']:
            for cell in row:
                cell.font = font
                cell.alignment = align

        for row in book['{}'.format(sheet)]['A12:Z19']:
            for cell in row:
                cell.alignment = align

    book.save(excel_path)

# 创建节点
def create_Node(element, text=None):
    elem = ET.Element(element)
    elem.text = text
    return elem

# 链接节点到根节点
def link_Node(root, element, text=None):
    """
    @param root: element的父节点
    @param element: 创建的element子节点
    @param text: element节点内容
    @return: 创建的子节点
    """
    if text != None:
        text = str(text)
    element = create_Node(element, text)
    root.append(element)
    return element

def compute_iou(bbox1, bbox2):
    """
    compute iou
    """
    bbox1xmin = bbox1[0]
    bbox1ymin = bbox1[1]
    bbox1xmax = bbox1[2]
    bbox1ymax = bbox1[3]
    bbox2xmin = bbox2[0]
    bbox2ymin = bbox2[1]
    bbox2xmax = bbox2[2]
    bbox2ymax = bbox2[3]

    area1 = (bbox1ymax - bbox1ymin) * (bbox1xmax - bbox1xmin)
    area2 = (bbox2ymax - bbox2ymin) * (bbox2xmax - bbox2xmin)

    bboxxmin = max(bbox1xmin, bbox2xmin)
    bboxxmax = min(bbox1xmax, bbox2xmax)
    bboxymin = max(bbox1ymin, bbox2ymin)
    bboxymax = min(bbox1ymax, bbox2ymax)
    if bboxxmin >= bboxxmax:
        return 0
    if bboxymin >= bboxymax:
        return 0

    area = (bboxymax - bboxymin) * (bboxxmax - bboxxmin)
    iou = area / (area1 + area2 - area)
    return iou

class Box:
    # x, y是左上角坐标
    def __init__(self, x, y, w, h, category=None, confidence=None):
        self.category = category
        self.confidence = confidence
        self.x = x
        self.y = y
        self.w = w
        self.h = h

    def get_area(self):
        return self.w * self.h

    def get_iou(self, box2):
        inter_area = calculate_inter_area(self, box2)
        return inter_area/(self.get_area()+box2.get_area()-inter_area)

def calculate_inter_area(box1, box2):
    '''
    :param box1: Box对象
    :param box2: Box对象
    :return: box1与box2的交面积
    '''
    left_x, left_y = max([box1.x, box2.x]), max([box1.y, box2.y])
    right_x, right_y = min([box1.x + box1.w, box2.x + box2.w]), min([box1.y + box1.h, box2.y + box2.h])
    height = right_y - left_y
    width = right_x - left_x
    area = height * width if height>0 and width>0 else 0
    return area

# -----以下代码用来创建文件夹-----
def make_dir(base_path):
    if not os.path.exists(base_path):
        os.mkdir(base_path)

def make_dir2(base_path, folders):
    if not os.path.exists(base_path):
        os.mkdir(base_path)
    for folder in folders:
        folder_path = os.path.join(base_path, folder)
        if not os.path.exists(folder_path):
            os.mkdir(folder_path)

def make_dir3(base_path, folders, subfolders):
    if not os.path.exists(base_path):
        os.mkdir(base_path)
    for folder in folders:
        folder_path = os.path.join(base_path, folder)
        if not os.path.exists(folder_path):
            os.mkdir(folder_path)
        for subfolder in subfolders:
            subfolder_path = os.path.join(folder_path, subfolder)
            if not os.path.exists(subfolder_path):
                os.mkdir(subfolder_path)
# -----以上代码用来创建文件夹-----

# 移动文件夹下所有指定尾缀文件到另一个文件夹
def move_specify_file(input_path, file_type, output_path):
    file_list = os.listdir(input_path)
    for file in file_list:
        if file.endswith(file_type):
            shutil.move(os.path.join(input_path, file), output_path)

# -----以下代码用来进行坐标转换-----
def extract_xys(axiss):
    '''
    :param axiss: xml中的坐标系父节点
    :return: list[x1,y1,...,xn,yn]
    '''
    return [float(axis.text) for axis in axiss if len(axis.text) > 5]

def points_to_xywh(obj):
    '''
    :param obj: labelme instance中待检测目标obj{}
    :return: box左上坐标+wh
    '''
    points = obj['points']
    shape_type = obj['shape_type']
    if shape_type == 'circle':
        center = [points[0][0], points[0][1]]
        radius = math.sqrt((points[1][0]-center[0])**2+(points[1][1]-center[1])**2)
        return [center[0]-radius-1, center[1]-radius-1, 2*radius+3, 2*radius+3]
    xs = [point[0] for point in points]
    ys = [point[1] for point in points]
    min_x, max_x = min(xs), max(xs)
    min_y, max_y = min(ys), max(ys)
    return [min_x-1, min_y-1, max_x-min_x+3, max_y-min_y+3]

def points_to_center(obj):
    '''
    :param obj: labelme instance中待检测目标obj{}
    :return: box中心坐标
    '''
    points = obj['points']
    shape_type = obj['shape_type']
    if shape_type == 'circle':
        return points[0][0], points[0][1]
    xs = [point[0] for point in points]
    ys = [point[1] for point in points]
    min_x, max_x = min(xs), max(xs)
    min_y, max_y = min(ys), max(ys)
    return (min_x+max_x)/2, (min_y+max_y)/2

def yolo_to_xywh(line):
    line = line.split(' ')
    x = float(line[1]) - float(line[3]) / 2
    y = float(line[2]) - float(line[4]) / 2
    w = float(line[3])
    h = float(line[4])
    category = int(line[0])
    return [x, y, w, h, category]
# -----以上代码用来进行坐标转换-----

# 字典对齐---以dic_a为基准，dic_b向dic_a对齐，补0
def dic_align(dic_a, dic_b):
    c = [i for i in dic_a if i not in dic_b]
    for i in c:
        dic_b[i] = 0
    return dic_b

def grid_search(a, b):
    """
    @param a: type->list [start, stop, N]
    @param b: type->list [start, stop, N]
    @return: two-dimensional array
    """
    # x, y = np.meshgrid(np.linspace(a[0], a[1], a[2]), np.linspace(b[0], b[1], b[2]))
    # cartesian_arr = np.array([x.ravel(),y.ravel().T])
    # return cartesian_arr.T
    x, y = np.meshgrid(np.linspace(a[0], a[1], a[2]), np.linspace(b[0], b[1], b[2]))
    cartesian_arr = np.array([x.ravel(),y.ravel().T])
    return np.round(cartesian_arr.T, 2)


################# 以下paul ###################

def instance_clean(instance):
    '''
    :param instance: labelme json instance
    :return: 将不良points进行清洗、更正
    '''
    for obj in instance['shapes']:
        points = obj['points']
        if obj['shape_type'] in ('line', 'linestrip'):
            # 排除标注小组的重复落点
            points_checked = [points[0]]
            for point in points:
                if point != points_checked[-1]:
                    points_checked.append(point)
            obj['points'] = points_checked
            points = points_checked
            # 排除标注小组的往返落点
            if len(points) >= 3:
                temp = get_angle((points[-3][0]-points[-2][0], points[-3][1]-points[-2][1]), (points[-1][0]-points[-2][0], points[-1][1]-points[-2][1]))
                if temp > -0.001 and temp < 0.001: del points[-1]
            if len(points) == 1: obj['shape_type'] = 'point'
            elif len(points) == 2: obj['shape_type'] = 'line'
            else: obj['shape_type'] = 'linestrip'

def instance_points_to_polygon(instance):
    '''
    :param instance: labelme json instance
    :return: 将instance['shapes']中points的标签，由rectangle和circle变为polygon，从而更好地进行crop和fill
    '''
    objs = instance['shapes']
    for obj in objs:
        shape_type = obj['shape_type']
        points = obj['points']
        if shape_type == 'rectangle':
            xs = [point[0] for point in points]
            ys = [point[1] for point in points]
            min_x, min_y = min(xs), min(ys)
            max_x, max_y = max(xs), max(ys)
            obj['points'] = [[min_x, min_y], [max_x, min_y], [max_x, max_y], [min_x, max_y]]
            obj['shape_type'] = 'polygon'
        elif shape_type == 'circle':
            center = [points[0][0], points[0][1]]
            radius = math.sqrt((points[1][0]-center[0])**2+(points[1][1]-center[1])**2)
            obj['points'] = []
            obj['shape_type'] = 'polygon'
            for i in range(0, 360, 10):
                obj['points'].append([center[0]+math.cos(math.pi*i/180)*radius, center[1]+math.sin(math.pi*i/180)*radius])

# -----以下代码求两线段交点坐标-----
class Point(object):
    def __init__(self, x=0, y=0):
        self.x = x
        self.y = y

class Line(object):
    # a=0, b=0, c=0
    def __init__(self, p1, p2):
        self.p1 = p1
        self.p2 = p2

def getLinePara(line):
    line.a =line.p1.y - line.p2.y
    line.b = line.p2.x - line.p1.x
    line.c = line.p1.x *line.p2.y - line.p2.x * line.p1.y

def getCrossPoint(l1, l2):
    getLinePara(l1)
    getLinePara(l2)
    d = l1.a * l2.b - l2.a * l1.b
    p=Point()
    if d == 0: return None
    p.x = (l1.b * l2.c - l2.b * l1.c)*1.0 / d
    p.y = (l1.c * l2.a - l2.c * l1.a)*1.0 / d
    return p

def get_cross_point(x1, y1, x2, y2, x3, y3, x4, y4):
    p1 = Point(x1, y1)
    p2 = Point(x2, y2)
    l1 = Line(p1, p2)
    p3 = Point(x3, y3)
    p4 = Point(x4, y4)
    l2 = Line(p3, p4)
    cp = getCrossPoint(l1, l2)
    if cp == None \
            or cp.x < min([x1, x2])-0.01 \
            or cp.x > max([x1, x2])+0.01 \
            or cp.y < min([y1, y2])-0.01 \
            or cp.y > max([y1, y2])+0.01 \
            or cp.x < min([x3, x4])-0.01 \
            or cp.x > max([x3, x4])+0.01 \
            or cp.y < min([y3, y4])-0.01 \
            or cp.y > max([y3, y4])+0.01:
        return None
    else:
        return [cp.x, cp.y]
# -----以上求两线段交点坐标-----

def points_to_coco_segmentation(obj, line_pixel):
    '''
    :param obj: labelme instance中待检测目标obj{}
    :param line_pixel: labelme中line、linestrip points的加宽像素值
    :return: coco segmentation[[x1,y1,x2,y2,...,xn,yn]]
    '''
    points = obj['points']
    shape_type = obj['shape_type']
    if shape_type == 'rectangle':
        xs = [point[0] for point in points]
        ys = [point[1] for point in points]
        min_x, min_y = min(xs), min(ys)
        max_x, max_y = max(xs), max(ys)
        result = [[min_x, min_y, max_x, min_y, max_x, max_y, min_x, max_y]]
    elif shape_type == 'circle':
        center = [points[0][0], points[0][1]]
        radius = math.sqrt((points[1][0]-center[0])**2+(points[1][1]-center[1])**2)
        temp = []
        for i in range(0, 360, 10):
            temp.append(center[0]+math.cos(math.pi*i/180)*radius)
            temp.append(center[1]+math.sin(math.pi*i/180)*radius)
        result = [temp]
    elif shape_type == 'line' or shape_type == 'linestrip':
        result = [line_pixel_widen(points, line_pixel)]
    else:
        result = [np.asarray(points).flatten().tolist()]
    return result

def line_pixel_widen(points, line_pixel):
    '''
    :param points: labelme中标签为line、linestrip的points
    :param line_pixel: 加宽的像素点
    :return: 返回coco中直线segmentation的坐标点
    '''
    line1 = []
    line2 = []
    for i, point in enumerate(points):
        belong_to_line1 = True
        if i == 0:
            vector2 = (points[i+1][0] - point[0], points[i+1][1] - point[1])
            angle_horiz = get_horiz_angle(vector2) + math.pi/2
            line1.append(perturbation_around_point(point, angle_horiz, line_pixel)[0])
            line2.append(perturbation_around_point(point, angle_horiz, line_pixel)[1])
            continue
        elif i == len(points)-1:
            vector1 = (points[i-1][0] - point[0], points[i-1][1] - point[1])
            angle_horiz = get_horiz_angle(vector1) + math.pi/2
            perturb_points = perturbation_around_point(point, angle_horiz, line_pixel)
            if get_cross_point(perturb_points[0][0], perturb_points[0][1], line1[-1][0], line1[-1][1], points[i-1][0], points[i-1][1], point[0], point[1]) != None:
                belong_to_line1 = False
        else:
            vector1 = (points[i-1][0] - point[0], points[i-1][1] - point[1])
            vector2 = (points[i+1][0] - point[0], points[i+1][1] - point[1])
            angle_horiz = get_mid_horiz_angle(vector1, vector2)
            radius = line_pixel/math.sin(get_angle(vector1, vector2)/2)
            perturb_points = perturbation_around_point(point, angle_horiz, radius)
            if get_cross_point(perturb_points[0][0], perturb_points[0][1], line1[-1][0], line1[-1][1], points[i-1][0], points[i-1][1], point[0], point[1]) != None:
                belong_to_line1 = False
        if belong_to_line1:
            line1.append(perturb_points[0])
            line2.append(perturb_points[1])
        else:
            line1.append(perturb_points[1])
            line2.append(perturb_points[0])
    # import matplotlib.pyplot as plt
    # line = [[points[i][0], points[i][1], points[i - 1][0], points[i - 1][1]] for i in range(1, len(points))]
    # plt.plot([p[0] for p in points], [p[1] for p in points], 'g-')
    # plt.plot([l[0] for l in line1], [l[1] for l in line1], 'b:')
    # plt.plot([l[0] for l in line2], [l[1] for l in line2], 'b:')
    # plt.show()
    return np.asarray(line1 + list(reversed(line2))).flatten().tolist()

def perturbation_around_point(point, angle, radius):
    '''
    :param point: 目标坐标点
    :param angle: 微扰角度
    :param radius: 微扰幅度
    :return: 返回一个坐标点的周围两个微扰点
    '''
    return [point[0] + radius * math.cos(angle), point[1] + radius * math.sin(angle)],\
           [point[0] - radius * math.cos(angle), point[1] - radius * math.sin(angle)]

def get_mid_horiz_angle(vector1, vector2):
    '''
    :param vector1: 向量1
    :param vector2: 向量2
    :return: 两个向量的中间向量与水平线(1,0)的夹角
    '''
    angle_horiz_1 = get_horiz_angle(vector1)
    angle_horiz_2 = get_horiz_angle(vector2)
    return (angle_horiz_1 + angle_horiz_2)/2

def get_horiz_angle(vector):
    '''
    :param vector: 一个向量
    :return: 该向量和水平线(1,0)的夹角(-180-180)
    '''
    angle_horiz = get_angle((1, 0), vector) if vector[1] > 0 else -get_angle((1, 0), vector)
    return angle_horiz

def get_angle(vector1, vector2):
    '''
    :param vector1: 向量1
    :param vector2: 向量2
    :return: 向量之间的夹角(0-180)
    '''
    inner_product = vector1[0]*vector2[0] + vector1[1]*vector2[1]
    cosin = inner_product/(math.sqrt((vector1[0]**2+vector1[1]**2)*(vector2[0]**2+vector2[1]**2)))
    return math.acos(cosin)

def crop_is_empty(instance, crop_size, iou_thres=0.2):
    '''
    :param instance: labelme json instance
    :param crop_size: crop范围[上，下，左，右]
    :return: bool值
    '''
    flag = True
    for obj in instance['shapes']:
        if obj_in_crop(obj, crop_size, iou_thres):
            flag = False
            break
    return flag

def obj_in_crop(obj, crop_size, iou_thres=0.2):
    '''
    :param points: labelme json中一个obj的points
    :param crop_size: crop范围[上，下，左，右]
    :param iou_thres: iou阈值
    :return: bool值
    '''
    crop_box = Box(crop_size[2], crop_size[0], crop_size[3] - crop_size[2], crop_size[1] - crop_size[0])
    x, y, w, h = points_to_xywh(obj)
    obj_box = Box(x, y, w, h)
    inter_area = calculate_inter_area(obj_box, crop_box)
    return inter_area != 0 and inter_area/obj_box.get_area() >= iou_thres

def point_in_crop(point, crop_size):
    '''
    :param point: labelme json中一个obj的points-point[x,y]
    :param crop_size: crop范围[上，下，左，右]
    :return: bool值
    '''
    return point[0] > crop_size[2] and \
           point[0] < crop_size[3] and \
           point[1] > crop_size[0] and \
           point[1] < crop_size[1]
################# 以上paul ###################


def get_crop_num(img_size, crop_size, overlap):
    '''
    :param img_size: img长或者宽
    :param crop_size: crop的边长
    :param overlap: 相邻框的交并比
    :return: 根据overlap和crop size计算滑框截取个数
    '''
    return math.ceil((img_size-crop_size)/((1-overlap)*crop_size)) + 1

def _random_crop(cx, cy, w, h, size, shift_x_left=0.75, shift_x_right=0.25, shift_y_up=0.75, shift_y_bottom=0.25):
    '''
    :param cx: 目标中心点x
    :param cy: 目标中心点y
    :param w: 图片width
    :param h: 图片height
    :param size: 截图的size
    :param shift_x_left: 截框左边框距离cx的最左随机范围（距离像素/size）
    :param shift_x_right: 截框左边框距离cx的最右随机范围（距离像素/size）
    :param shift_y_up: 截框上边框距离cy的最上随机范围（距离像素/size）
    :param shift_y_bottom: 截框上边框距离cy的最下随机范围（距离像素/size）
    :return: 返回随机截图框
    '''
    # 截框左边框、上边框距离目标中心点的offset
    ofx, ofy = random.randint(int(size*shift_x_right), int(size*shift_x_left)), random.randint(int(size*shift_y_bottom), int(size*shift_y_up))
    cx, cy = int(cx), int(cy)
    fill_size = [0, 0, 0, 0]
    if size > h:
        up, bottom = 0, h
        fill_size[0], fill_size[1] = (size-h)//2, size-h-(size-h)//2
    elif cy-ofy < 0:
        up, bottom = 0, size
    elif cy-ofy+size > h:
        up, bottom = h-size, h
    else:
        up, bottom = cy-ofy, cy-ofy+size
    if size > w:
        left, right = 0, w
        fill_size[2], fill_size[3] = (size-w)//2, size-w-(size-w)//2
    elif cx-ofx < 0:
        left, right = 0, size
    elif cx-ofx+size > w:
        left, right = w-size, w
    else:
        left, right = cx-ofx, cx-ofx+size
    return [up, bottom, left, right], fill_size

# 根据过检和漏失的增强截图策略
def aug_crop_strategy(img, instance, size, additional_info):
    # 过检增强倍数
    precision_aug = 30
    # 漏失增强倍数
    recall_aug = 0
    # 错误、难样本增强倍数
    hard_aug = 0
    crop_strategies = []
    w = instance['imageWidth']
    h = instance['imageHeight']
    for obj in instance['shapes']:
        label = obj['label']
        # obj中心坐标位置
        cx, cy = points_to_center(obj)
        if label.startswith('guojian'):
            for i in range(precision_aug):
                crop_strategies.append(_random_crop(cx, cy, w, h, size))
        elif label.startswith('loushi'):
            for i in range(recall_aug):
                crop_strategies.append(_random_crop(cx, cy, w, h, size))
        elif label.startswith('hard') or label.startswith('cuowu'):
            for i in range(hard_aug):
                crop_strategies.append(_random_crop(cx, cy, w, h, size))
    return crop_strategies

# 根据过检和漏失的动态增强截图策略
def dynamic_aug_crop_strategy(img, instance, size, additional_info):
    crop_strategies = []
    w = instance['imageWidth']
    h = instance['imageHeight']
    for obj in instance['shapes']:
        label = obj['label']
        # 正常标签，不需要增强，则continue
        if '_' not in label: continue
        # 在label字符串中，p的起始、结束位置
        start, end = label.index('_')+1, label.index('_', label.index('_')+1)
        # obj中心坐标位置
        cx, cy = points_to_center(obj)
        if label.startswith('guojian'):
            p = float(label[start: end])
            for i in range(_dynamic_function(p, False)):
                crop_strategies.append(_random_crop(cx, cy, w, h, size))
        elif label.startswith('cuowu') or label.startswith('loushi'):
            for i in range(_dynamic_function(0.1)):
                crop_strategies.append(_random_crop(cx, cy, w, h, size))
        elif label.startswith('hard'):
            p = float(label[start: end])
            for i in range(_dynamic_function(p)):
                crop_strategies.append(_random_crop(cx, cy, w, h, size))
    return crop_strategies

def _dynamic_function(p, pos=True):
    '''
    :param p: 概率
    :param pos: 正标签or负标签
    :return: 动态增强倍数
    '''
    return round(10*(1-p)**2) if pos else 1

# 检测特定标签的截图策略
def check_crop_strategy(img, instance, size, additional_info):
    # 是否目标居中
    centerness = True
    # 需要查看的labels
    check_list = additional_info['check_list']
    crop_strategies = []
    w = instance['imageWidth']
    h = instance['imageHeight']
    for obj in instance['shapes']:
        label = obj['label']
        is_target = False
        for target in check_list:
            if target in label:
                is_target = True
                break
        if not is_target: continue
        cx, cy = points_to_center(obj)
        if centerness:
            crop_strategies.append(_random_crop(cx, cy, w, h, size, 0.5, 0.5, 0.5, 0.5))
        else:
            crop_strategies.append(_random_crop(cx, cy, w, h, size))
    return crop_strategies

# 聚类截图策略
def clustering_crop_strategy(img, instance, size, additional_info):
    crop_strategies = []
    added = []  # 用来存放截取过的obj
    h, w = img.shape[0], img.shape[1]
    # w = instance['imageWidth']
    # h = instance['imageHeight']
    objs = instance['shapes']
    num = len(objs)
    for i, obj in enumerate(objs):
        # 如果obj被截取过，continue
        if obj in added: continue
        # 当前聚类的外边框
        current_box = Box(*points_to_xywh(obj))
        # 开始搜寻聚类的objs
        for j in range(i+1, num):
            # 下一个obj
            next_obj = objs[j]
            # 如果下一个obj被截取过，continue
            if next_obj in added: continue
            next_box = Box(*points_to_xywh(next_obj))
            # 将下一个obj融合进当前的聚类的外边框
            combine_box = _combine_boxes(current_box, next_box)
            # 如果下一个obj不适合聚类，continue
            if combine_box.w > size or combine_box.h > size: continue
            # 聚类完成，更新当前的聚类的外边框
            current_box = combine_box
            # 将下一个obj放入added列表
            added.append(next_obj)
        if current_box.w < size and current_box.h < size:
            crop_strategies.append(_random_crop(current_box.x+current_box.w/2, current_box.y+current_box.h/2, w, h, size, (size-current_box.w/2)/size, current_box.w/2/size, (size-current_box.h/2)/size, current_box.h/2/size))
        else:
            crop_strategies.append(_random_crop(current_box.x+current_box.w/2, current_box.y+current_box.h/2, w, h, size, 0.5, 0.5, 0.5, 0.5))
    return crop_strategies

def _combine_boxes(box1, box2):
    '''
    :param box1:
    :param box2:
    :return: 返回两个box的合并box
    '''
    xmin = min(box1.x, box2.x)
    ymin = min(box1.y, box2.y)
    xmax = max(box1.x+box1.w, box2.x+box2.w)
    ymax = max(box1.y+box1.h, box2.y+box2.h)
    return Box(xmin, ymin, xmax-xmin, ymax-ymin)


def imgs_crop_and_fill(img_folder_path, crop_strategy, img_size, output_path, empty_check, num_worker, iou_thres, additional_info):
    '''
    :param img_folder_path: 图片文件夹绝对路径
    :param crop_strategy: 截取策略
    :param img_size: 截图的size
    :param empty_check: 是否检查空截框
    :param num_worker: 线程数
    :param output_path: 保存路径
    :param iou_thres: 截图时iou阈值
    :param additional_info: information for crop_strategy method
    :return: 整个文件夹下的图片截图和填充
    '''
    thread_pool = ThreadPoolExecutor(max_workers=num_worker)
    print('Thread Pool is created!')
    # 检查输出路径文件夹是否存在
    if not os.path.exists(output_path): os.makedirs(output_path)
    for img_name in os.listdir(img_folder_path):
        img_file_path = os.path.join(img_folder_path, img_name)
        # 过滤文件夹和非图片文件
        if not os.path.isfile(img_file_path) or img_name[img_name.rindex('.')+1:] not in IMG_TYPES: continue
        thread_pool.submit(img_crop_and_fill, img_file_path, crop_strategy, img_size, output_path, empty_check, iou_thres, additional_info)
    thread_pool.shutdown(wait=True)

# 单张图片的截图和填充，用来作为多线程的输入方法
def img_crop_and_fill(img_file_path, crop_strategy, img_size, output_path, empty_check, iou_thres, additional_info):
    img = cv2.imread(img_file_path)
    try:
        instance = json_to_instance(img_file_path[:img_file_path.rindex('.')]+'.json')
    except FileNotFoundError:
        # 有些图片没有对应的json文件，或表示该图片无目标
        print('\033[1;33m%s has no json file...\033[0m' % (img_file_path))
        instance = create_empty_json_instance(img_file_path)
    instance_points_to_polygon(instance)
    crops = crop_strategy(img, instance, img_size, additional_info)
    for crop in crops:
        crop_size, fill_size = crop[0], crop[1]
        try:
            init_crop_and_fill(img, instance, crop_size, fill_size, output_path, empty_check, iou_thres)
        except Exception as e:
            # 有些图片crop报错需要log
            print('\033[1;31m%s fails in cropping %s, due to %s.\033[0m' % (img_file_path, crop_size, e.with_traceback()))

# 单张图片的截图和填充，要求输入已经读取的img和json instance
def init_crop_and_fill(img, instance, crop_size, fill_size, output_path, empty_check, iou_thres):
    if empty_check and crop_is_empty(instance, crop_size, iou_thres): return
    # 新图片名、json文件名，新图片路径，json文件路径
    offset_x, offset_y = crop_size[2]-fill_size[2], crop_size[0]-fill_size[0]
    img_new_name = instance['imagePath'].replace('.', '_%d_%d.' % (offset_x, offset_y))
    json_new_name = img_new_name[:img_new_name.rindex('.')] + '.json'
    img_new_path = os.path.join(output_path, img_new_name)
    json_new_path = os.path.join(output_path, json_new_name)
    instance_new = {'version': '1.0', 'imageData': None,
                    'imageWidth': crop_size[3] - crop_size[2] + fill_size[2] + fill_size[3],
                    'imageHeight': crop_size[1] - crop_size[0] + fill_size[0] + fill_size[1],
                    'imageDepth': img.shape[2],
                    'imagePath': img_new_name,
                    'shapes': copy.deepcopy(instance['shapes'])}
    # 先截图后填充
    img_crop = img[crop_size[0]: crop_size[1], crop_size[2]: crop_size[3]]
    update_objs_in_crop(instance_new, crop_size, iou_thres)
    img_new = cv2.copyMakeBorder(img_crop, fill_size[0], fill_size[1], fill_size[2], fill_size[3], cv2.BORDER_REPLICATE)
    cv2.imwrite(img_new_path, img_new)
    for obj in instance_new['shapes']:
        for point in obj['points']:
            point[0] -= offset_x
            point[1] -= offset_y
    instance_to_json(instance_new, json_new_path)
    print(img_new_name, ' is done!')

# instance中的shapes字段为原instance中shapes字段的深拷贝
# 此方法更新instance中的shapes字段
def update_objs_in_crop(instance, crop_size, iou_thres=0.2):
    shapes_new = []
    shapes = instance['shapes']
    # 遍历shapes中的目标objs
    for obj in shapes:
        points = obj['points']
        shape_type = obj['shape_type']
        # 目标不在crop区域，continue
        if not obj_in_crop(obj, crop_size, iou_thres):
            continue
        # 目标在crop区域，开始更新坐标
        points_new = []
        # 四条截边
        bounds = [[crop_size[2], crop_size[0], crop_size[3], crop_size[0]],  # (xmin, ymin, xmax, ymin)
                  [crop_size[3], crop_size[0], crop_size[3], crop_size[1]],  # (xmax, ymin, xmax, ymax)
                  [crop_size[3], crop_size[1], crop_size[2], crop_size[1]],  # (xmax, ymax, xmin, ymax)
                  [crop_size[2], crop_size[1], crop_size[2], crop_size[0]]]  # (xmin, ymax, xmin, ymin)
        # four_points = [[crop_size[2], crop_size[0]], [crop_size[3], crop_size[0]], [crop_size[3], crop_size[1]], [crop_size[2], crop_size[1]]]
        # xywh = points_to_xywh(obj)
        for i, point in enumerate(points):
            if point_in_crop(point, crop_size):
                if (i != 0 or shape_type == 'polygon') and (not point_in_crop(points[i-1], crop_size)):
                    for bound in bounds:
                        cross_point = get_cross_point(point[0], point[1], points[i-1][0], points[i-1][1], *bound)
                        if cross_point != None:
                            points_new.append(cross_point)
                            break
                points_new.append(point)
            elif (i != 0 or shape_type == 'polygon') and point_in_crop(points[i-1], crop_size):
                for bound in bounds:
                    cross_point = get_cross_point(point[0], point[1], points[i-1][0], points[i-1][1], *bound)
                    if cross_point != None:
                        points_new.append(cross_point)
                        break
            elif (i != 0 or shape_type == 'polygon') and (not point_in_crop(points[i-1], crop_size)):
                temp = []
                for bound in bounds:
                    cross_point = get_cross_point(point[0], point[1], points[i-1][0], points[i-1][1], *bound)
                    if cross_point != None:
                        temp.append(cross_point)
                if len(temp) == 0: continue
                if (temp[0][0]-point[0])**2+(temp[0][1]-point[1])**2 > (temp[1][0]-point[0])**2+(temp[1][1]-point[1])**2:
                    points_new.append(temp[0])
                    points_new.append(temp[1])
                else:
                    points_new.append(temp[1])
                    points_new.append(temp[0])
        obj['points'] = points_new
        shapes_new.append(obj)
    instance['shapes'] = shapes_new

def obj_crop(target_folder_path, output_path):
    '''
    :param target_folder_path: labelme folder path
    :return: only crop obj area
    '''
    # 检查输出路径文件夹是否存在
    if not os.path.exists(output_path): os.makedirs(output_path)
    for img_file in os.listdir(target_folder_path):
        img_file_path =os.path.join(target_folder_path, img_file)
        # 过滤文件夹和非图片文件
        if not os.path.isfile(img_file_path) or img_file[img_file.rindex('.')+1:] not in IMG_TYPES: continue
        img = cv2.imread(img_file_path)
        instance = json_to_instance(os.path.join(target_folder_path, img_file[:img_file.rindex('.')]+'.json'))
        for obj in instance['shapes']:
            x, y, w, h = points_to_xywh(obj)
            img_new = img[y:y+h, x:x+w]
            cv2.imwrite(os.path.join(output_path, img_file.replace('.', '_%d_%d.'%(x, y))), img_new)
            print('Image %s obj %s is cropped!' % (img_file, obj['label']))
    print('Finished!')

def img_crop(input_dir, crop_size):
    imgs_crop_and_fill(img_folder_path=input_dir,  # img and json should be put in one folder
                       # 自定义的截图策略
                       crop_strategy=clustering_crop_strategy,  # aug_crop_strategy, clustering_crop_strategy
                       # 截图尺寸
                       img_size=crop_size,
                       # 截图输出路径
                       output_path=input_dir + '_crop'+ str(crop_size),  # Automatically create output folders
                       # 自动滤去不含检测目标的截图框
                       empty_check=False,
                       # 多线程数
                       num_worker=8,
                       # 被截断的检测目标的面积比阈值，低于阈值将不计入截图框中
                       iou_thres=0.2,
                       # information for crop_strategy method
                       additional_info={'check_list': ['jiaobuliang', 'jiaobuliang-1', 'jiaobuliang-2']})

if __name__ == '__main__':
    # 在这里定义自己的crop和fill的strategy
    # crop_strategy方法：根据img对象和json instance，给出crop和fill的范围，遵循上下左右
    # return [[crop1, fill1], [crop2, fill2],...]
    # crop: [top, bottom, left, right]   fill: [top, bottom, left, right]
    # img为opencv读取的图片对象，instance为json对象
    # 内置的crop_strategy方法在img_slice_utils.py中
    def define_my_crop_strategy(img, instance, img_size, additional_info):
        h, w = instance['imageHeight'], instance['imageWidth']
        pad = (1120 - w)//2
        return [[[0, h, 0, w], [0, 0, pad, pad]]]
    # 截图并保存
    input_dir = '/home/jerry/Desktop/garbage/1_test_demo/bbaug-demo'
    crop_size = 768
    img_crop(input_dir, crop_size)
































