import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from .conversionhours import conversionHours

class conversion:
    def __init__(self, file_name):
        self.file_path = file_name

    def convert(self):
        if os.path.exists("converted.xlsx"):
            os.remove("converted.xlsx")
        indexs = []
        pd_file = pd.read_excel(self.file_path)
        d1 = pd_file["Unnamed: 8"]
        d2 = pd_file["Unnamed: 9"]
        d3 = pd_file["Unnamed: 10"]
        d4 = pd_file["Unnamed: 11"]

        result = [d1, d2, d3, d4]


        for i in range(len(result)):
            w1 = conversionHours(result[i])
            for i in range(len(w1)):
                if w1[i] == "NO":
                    w1[i] = 0

            d = pd.DataFrame(w1)
            pd_file = pd.concat([pd_file,d], axis=1)

        pd_file.columns = ["Ticket Number", "Title","Ticket Category","Ticket Raised For","Employee Job Title"	,"Requested On","Priority" ,"Assigned To" , "Expected First Response Time",	"Actual First Response Time", "Expected Resolution Time" , "Actual Resolution Time", "Ticket Status", "Last Updated", "Closed On", "Closed By", "ExpectedFirstResponseTime_converted", "ActualFirstResponseTime_converted", "ExpectedResolutionTime_converted", "ActualResolutionTime_converted"]
        pd_file.to_excel("converted.xlsx")

        wb = openpyxl.load_workbook("converted.xlsx")
        ws = wb["Sheet1"]

        c1 = pd_file["ExpectedFirstResponseTime_converted"]
        c2 = pd_file["ActualFirstResponseTime_converted"]
        c3 = pd_file["ExpectedResolutionTime_converted"]
        c4 = pd_file["ActualResolutionTime_converted"]


        for j in range(len(c2)):
            if c1[j] < c2[j] and c1[j] != 0:
                ws[f"S{j+2}"].fill = PatternFill(patternType='solid',fgColor='FC2C03')
                wb.save("converted.xlsx")

        for z in range(len(c2)):
            if c3[z] < c4[z] and c3[z] != 0:
                ws[f"U{z+2}"].fill = PatternFill(patternType='solid',fgColor='FC2C03')
                wb.save("converted.xlsx")